"""Service Library torque verification helpers."""

import re
import time
from difflib import SequenceMatcher
from functools import lru_cache
from io import StringIO
from urllib.parse import urlparse, parse_qs
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
import requests

import config
import parser


CONFIG_LEVEL = "YEAR_MODEL_ENGINE"
REQUEST_TIMEOUT = 30
SHORTCUT_WORKBOOK_PATH = Path(__file__).with_name("Accroname_DT-26_enriched.xlsx")
MAX_DESCRIPTION_VARIANTS = 64
MODULE_EXPANSION_MIN_SCORE = 0.85
TORQUE_CONVERSIONS_TO_NM = {
    "nm": 1.0,
    "n_cm": 0.01,
    "n_mm": 0.001,
    "ft_lb": 1.3558179483314,
    "in_lb": 0.1129848290276167,
    "kgf_m": 9.80665,
    "kgf_cm": 0.0980665,
}
TORQUE_UNIT_PATTERNS = (
    ("kgf_cm", r"(?:kgf|kg)\s*[- ]?\s*cm\b"),
    ("kgf_m", r"(?:kgf|kg)\s*[- ]?\s*m\b"),
    ("ft_lb", r"(?:ft|foot|feet)\.?\s*[- ]?\s*(?:lb|lbs|lbf)\.?"),
    ("ft_lb", r"(?:lb|lbs|lbf)\.?\s*[- ]?\s*(?:ft|foot|feet)\.?"),
    ("in_lb", r"(?:in|inch|inches)\.?\s*[- ]?\s*(?:lb|lbs|lbf)\.?"),
    ("in_lb", r"(?:lb|lbs|lbf)\.?\s*[- ]?\s*(?:in|inch|inches)\.?"),
    ("n_cm", r"n\s*[.\- ]?\s*cm\b"),
    ("n_mm", r"n\s*[.\- ]?\s*mm\b"),
    ("nm", r"n\s*[.\- ]?\s*m\b"),
)
NUMBER_PATTERN = r"[-+]?\d+(?:\.\d+)?"
TORQUE_TABLE_UNIT_COLUMNS = {
    "nm": "nm",
    "n m": "nm",
    "newton meter": "nm",
    "newton meters": "nm",
    "ft lb": "ft_lb",
    "ft lbs": "ft_lb",
    "foot pound": "ft_lb",
    "foot pounds": "ft_lb",
    "lb ft": "ft_lb",
    "lbs ft": "ft_lb",
    "in lb": "in_lb",
    "in lbs": "in_lb",
    "inch pound": "in_lb",
    "inch pounds": "in_lb",
    "lb in": "in_lb",
    "lbs in": "in_lb",
    "torque": "nm",
    "torque nm": "nm",
    "value": "nm",
}
TORQUE_UNIT_DISPLAY = {
    "nm": "N m",
    "ft_lb": "ft lb",
    "in_lb": "in lb",
    "n_cm": "N cm",
    "n_mm": "N mm",
    "kgf_m": "kgf m",
    "kgf_cm": "kgf cm",
}

_shortcut_cache_signature: Optional[Tuple[int, int]] = None
_shortcut_cache: Dict[Tuple[str, ...], Tuple[Dict[str, str], ...]] = {}

# These keep the core torque matcher useful if the external workbook is missing.
# The workbook remains the primary source and can contain multiple meanings.
CORE_DESCRIPTION_SHORTCUTS = {
    "rr": ("rear",),
    "fr": ("front",),
    "frt": ("front",),
    "lh": ("left hand",),
    "rh": ("right hand",),
    "lwr": ("lower",),
    "upr": ("upper",),
    "ctrl": ("control",),
    "assy": ("assembly",),
    "brkt": ("bracket",),
    "mtg": ("mounting",),
    "brg": ("bearing",),
    "cyl": ("cylinder",),
    "eng": ("engine",),
    "trans": ("transmission",),
    "diff": ("differential",),
    "mt": ("mount", "mounting", "manual transmission"),
    "asm": ("assembly",),
    "damp": ("damper",),
    "plg": ("plug",),
    "abs": ("absorber",),
    "em": ("engine mount",),
}


def _headers(accept: str = "application/json, text/javascript, */*; q=0.01") -> Dict[str, str]:
    headers = config.get_headers({"Accept": accept})
    cookies = config.get_cookies()
    if cookies:
        headers["Cookie"] = cookies
    return headers


def _get_json(path: str, params: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    response = requests.get(
        f"{config.BASE_URL}{path}",
        headers=_headers(),
        params=params,
        timeout=REQUEST_TIMEOUT,
    )
    response.raise_for_status()
    return response.json()


def _get_text(path: str, params: Optional[Dict[str, Any]] = None) -> str:
    response = requests.get(
        f"{config.BASE_URL}{path}",
        headers=_headers("*/*"),
        params=params,
        timeout=REQUEST_TIMEOUT,
    )
    response.raise_for_status()
    return response.text


def _get_torque_content(leaf: Dict[str, str], model_version_engine_id: str) -> str:
    try:
        return _get_text(f"/connect/api/content/raw/{leaf['content_link_id']}")
    except requests.HTTPError as exc:
        response = exc.response
        if response is None or response.status_code not in {400, 404}:
            raise

    context_paths = [
        f"/connect/api/content/raw/{leaf['content_link_id']}/{CONFIG_LEVEL}/{model_version_engine_id}",
        f"/connect/api/discipline/content/raw/{leaf['content_link_id']}/{CONFIG_LEVEL}/{model_version_engine_id}",
    ]
    info_codes = [leaf.get("info_code") or "undefined", "undefined", "", None]
    last_error = None
    seen = set()
    for context_path in context_paths:
        for info_code in info_codes:
            marker = (context_path, "<missing>" if info_code is None else info_code)
            if marker in seen:
                continue
            seen.add(marker)

            params = {"locale": config.MODEL_LOCALE, "container": "main"}
            if info_code is not None:
                params["infoCode"] = info_code
            auth_token = config.get_auth_token()
            if auth_token:
                params["X-Auth-Token"] = auth_token

            try:
                html = _get_text(context_path, params=params)
                # If the returned shell links to a raw content URL with null placeholders,
                # replace the nulls with the correct config level and model_version_engine_id
                if "connect/api/content/raw" in html and "/null/null" in html:
                    corrected_path = f"/connect/api/content/raw/{leaf['content_link_id']}/{CONFIG_LEVEL}/{model_version_engine_id}"
                    try:
                        return _get_text(corrected_path, params={"locale": config.MODEL_LOCALE, "container": "main", "infoCode": info_code or "undefined", "X-Auth-Token": auth_token} if auth_token else {"locale": config.MODEL_LOCALE, "container": "main", "infoCode": info_code or "undefined"})
                    except requests.HTTPError:
                        # fall through to return the original html if corrected fetch fails
                        pass
                return html
            except requests.HTTPError as exc:
                last_error = exc
                response = exc.response
                if response is None or response.status_code not in {400, 404}:
                    raise

    if last_error:
        raise last_error
    return _get_text(context_paths[0])


def _request_status(path: str, params: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    try:
        text = _get_text(path, params=params)
        return {"ok": True, "status": "OK", "detail": f"{len(text)} chars"}
    except requests.RequestException as exc:
        response = exc.response
        if response is None:
            return {"ok": False, "status": exc.__class__.__name__, "detail": str(exc)}
        return {"ok": False, "status": str(response.status_code), "detail": response.reason}


def _get_content_metadata(leaf: Dict[str, str], model_version_engine_id: str) -> Dict[str, Any]:
    return _get_json(
        f"/connect/api/metadata/{leaf['content_link_id']}/{CONFIG_LEVEL}/{model_version_engine_id}",
        params={"locale": config.MODEL_LOCALE, "infoCode": leaf.get("info_code") or "undefined"},
    )


def _plugin_iframe_params(
    leaf: Dict[str, str],
    model_version_engine_id: str,
    metadata: Dict[str, Any],
) -> Dict[str, str]:
    auth_token = config.get_auth_token()
    return {
        "contentLinkId": leaf["content_link_id"],
        "infoCode": leaf.get("info_code") or "undefined",
        "externalContentKey": metadata.get("externalContentKey") or "",
        "infoTypeCode": metadata.get("infoTypeCode") or "",
        "configLevel": CONFIG_LEVEL,
        "configId": model_version_engine_id,
        "contentType": metadata.get("contentType") or "",
        "contentPluginType": metadata.get("contentPluginType") or "",
        "X-Auth-Token": auth_token,
        "height": "800",
        "width": "1200",
        "container": "main",
        "anchorTag": "",
        "serviceActionId": metadata.get("serviceActionId") or "",
        "vehicleIssueId": metadata.get("vehicleIssueId") or "",
        "acknowledgementServiceActionId": metadata.get("acknowledgementServiceActionId") or "",
        "resolutionServiceActionId": metadata.get("resolutionServiceActionId") or "",
        "feedbackSubmissionTypeList": metadata.get("feedbackSubmissionTypeList") or "",
    }


def diagnose_torque_api(
    model_year: int,
    vehicle_family: str,
    engine_code: str,
) -> Dict[str, Any]:
    """Run a safe end-to-end torque API diagnostic without exposing credentials."""
    steps = []

    def add(step: str, ok: bool, detail: str) -> None:
        steps.append({"Step": step, "Result": "OK" if ok else "Failed", "Details": detail})

    vehicles = _find_vehicle_versions(model_year, vehicle_family)
    add("Vehicle lookup", bool(vehicles), f"{len(vehicles)} vehicle match(es)")
    if not vehicles:
        return {"steps": steps, "raw_content_attempts": []}

    engine_targets, checked_engines = _build_engine_targets(vehicles, engine_code)
    add(
        "Engine lookup",
        bool(engine_targets),
        f"{len(engine_targets)} engine target(s), {len(checked_engines)} vehicle(s) checked",
    )
    if not engine_targets:
        return {"steps": steps, "raw_content_attempts": []}

    target = engine_targets[0]
    engine = target["engine"]
    service_book = _get_service_book(engine["model_version_engine_id"])
    add("Service book lookup", bool(service_book), service_book.get("modelVersionBookId", "") if service_book else "")
    if not service_book:
        return {"steps": steps, "raw_content_attempts": []}

    toc = _get_json(
        f"/connect/api/toc/{service_book['modelVersionBookId']}/{CONFIG_LEVEL}/{engine['model_version_engine_id']}",
        params={"locale": config.MODEL_LOCALE, "nocache": str(int(time.time() * 1000))},
    )
    leaves = _rank_leaves(_collect_torque_leaves(toc), "")
    add("Torque TOC lookup", bool(leaves), f"{len(leaves)} torque page(s)")
    if not leaves:
        return {"steps": steps, "raw_content_attempts": []}

    leaf = leaves[0]
    metadata = _get_content_metadata(leaf, engine["model_version_engine_id"])
    access_ok = bool(
        metadata.get("subscription", {}).get("allowsAccess", True)
        and metadata.get("restriction", {}).get("userHasAllRestrictions", True)
    )
    add(
        "Content metadata lookup",
        access_ok,
        (
            f"{metadata.get('contentType', '')}/{metadata.get('contentPluginType', '')}; "
            f"external key {metadata.get('externalContentKey', '')}"
        ),
    )

    plugin_status = _request_status(
        "/web/secure/api/plugin/iframe",
        params=_plugin_iframe_params(leaf, engine["model_version_engine_id"], metadata),
    )
    add(
        "Plugin iframe lookup",
        plugin_status["ok"],
        f"{plugin_status['status']}; {plugin_status['detail']}",
    )

    context_paths = [
        ("Content context raw", f"/connect/api/content/raw/{leaf['content_link_id']}/{CONFIG_LEVEL}/{engine['model_version_engine_id']}"),
        ("Discipline context raw", f"/connect/api/discipline/content/raw/{leaf['content_link_id']}/{CONFIG_LEVEL}/{engine['model_version_engine_id']}"),
    ]
    auth_token = config.get_auth_token()
    raw_attempts = []
    attempts = [
        ("Simple raw content", f"/connect/api/content/raw/{leaf['content_link_id']}", None),
    ]
    for path_label, context_path in context_paths:
        for info_label, info_code in (
            ("TOC infoCode", leaf.get("info_code") or "undefined"),
            ("undefined infoCode", "undefined"),
            ("blank infoCode", ""),
            ("no infoCode", None),
        ):
            params = {"locale": config.MODEL_LOCALE, "container": "main"}
            if info_code is not None:
                params["infoCode"] = info_code
            if auth_token:
                params["X-Auth-Token"] = auth_token
            attempts.append((f"{path_label}: {info_label}", context_path, params))

    for label, path, params in attempts:
        status = _request_status(path, params=params)
        raw_attempts.append(
            {
                "Attempt": label,
                "Result": "OK" if status["ok"] else "Failed",
                "Status": status["status"],
                "Details": status["detail"],
            }
        )

    add(
        "Raw content lookup",
        any(attempt["Result"] == "OK" for attempt in raw_attempts),
        f"{sum(attempt['Result'] == 'OK' for attempt in raw_attempts)} successful raw content variant(s)",
    )
    return {"steps": steps, "raw_content_attempts": raw_attempts}


def _walk(value: Any) -> Iterable[Any]:
    yield value
    if isinstance(value, dict):
        for child in value.values():
            yield from _walk(child)
    elif isinstance(value, list):
        for child in value:
            yield from _walk(child)


def _clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _match_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def _shortcut_workbook_signature() -> Optional[Tuple[int, int]]:
    try:
        stat = SHORTCUT_WORKBOOK_PATH.stat()
    except OSError:
        return None
    return stat.st_mtime_ns, stat.st_size


def _description_shortcuts() -> Dict[Tuple[str, ...], Tuple[Dict[str, str], ...]]:
    global _shortcut_cache_signature, _shortcut_cache

    signature = _shortcut_workbook_signature()
    if signature == _shortcut_cache_signature and _shortcut_cache:
        return _shortcut_cache

    shortcuts: Dict[Tuple[str, ...], List[Dict[str, str]]] = {}

    if signature is not None:
        workbook = None
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(SHORTCUT_WORKBOOK_PATH, read_only=True, data_only=True)
            worksheet = workbook["List"]
            rows = worksheet.iter_rows(values_only=True)
            headers = next(rows, ())
            category_index = next(
                (
                    index
                    for index, header in enumerate(headers)
                    if _match_key(header) in {"category", "type"}
                ),
                None,
            )
            for row in rows:
                shortcut = row[0] if len(row) > 0 else ""
                description = row[1] if len(row) > 1 else ""
                shortcut_key = tuple(_match_key(shortcut).split())
                description_key = _match_key(description)
                if not shortcut_key or not description_key:
                    continue
                meanings = shortcuts.setdefault(shortcut_key, [])
                category = (
                    _match_key(row[category_index])
                    if category_index is not None and len(row) > category_index
                    else "mechanical"
                )
                if category not in {"module", "mechanical"}:
                    category = "mechanical"
                existing = next((item for item in meanings if item["text"] == description_key), None)
                if existing:
                    if category == "mechanical":
                        existing["category"] = category
                else:
                    meanings.append({"text": description_key, "category": category})
        except (OSError, ValueError, KeyError):
            pass
        finally:
            if workbook is not None:
                workbook.close()

    for shortcut, descriptions in CORE_DESCRIPTION_SHORTCUTS.items():
        shortcut_key = tuple(_match_key(shortcut).split())
        meanings = shortcuts.setdefault(shortcut_key, [])
        for description in descriptions:
            description_key = _match_key(description)
            existing = next((item for item in meanings if item["text"] == description_key), None)
            if existing:
                existing["category"] = "mechanical"
            elif description_key:
                meanings.append({"text": description_key, "category": "mechanical"})

    _shortcut_cache_signature = signature
    _shortcut_cache = {shortcut: tuple(meanings) for shortcut, meanings in shortcuts.items()}
    _description_variants_cached.cache_clear()
    return _shortcut_cache


@lru_cache(maxsize=256)
def _description_variants_cached(
    value: str,
    signature: Optional[Tuple[int, int]],
) -> List[Dict[str, Any]]:
    original_key = _match_key(value)
    if not original_key:
        return [{"text": "", "expansions": []}]

    tokens = original_key.split()
    shortcuts = _description_shortcuts()
    variants: List[Dict[str, Any]] = []

    def visit(index: int, output: List[str], expansions: List[Dict[str, Any]]) -> None:
        if len(variants) >= MAX_DESCRIPTION_VARIANTS:
            return
        if index >= len(tokens):
            variants.append({"text": " ".join(output), "expansions": expansions})
            return

        matches = [
            shortcut
            for shortcut in shortcuts
            if len(shortcut) <= len(tokens) - index
            and tuple(tokens[index:index + len(shortcut)]) == shortcut
        ]
        if not matches:
            visit(index + 1, output + [tokens[index]], expansions)
            return

        shortcut = max(matches, key=len)
        shortcut_text = " ".join(shortcut)
        meanings = shortcuts[shortcut]
        is_ambiguous = len(meanings) > 1
        for meaning_item in meanings:
            meaning = meaning_item["text"]
            visit(
                index + len(shortcut),
                output + meaning.split(),
                expansions
                + [
                    {
                        "shortcut": shortcut_text.upper(),
                        "meaning": meaning,
                        "ambiguous": is_ambiguous,
                        "category": meaning_item["category"],
                    }
                ],
            )

    visit(0, [], [])
    variants.append({"text": original_key, "expansions": []})

    unique = []
    seen = set()
    for variant in variants:
        marker = (
            variant["text"],
            tuple((item["shortcut"], item["meaning"]) for item in variant["expansions"]),
        )
        if marker in seen:
            continue
        seen.add(marker)
        unique.append(variant)
    return unique


def _description_variants(value: str) -> List[Dict[str, Any]]:
    _description_shortcuts()
    return _description_variants_cached(value, _shortcut_cache_signature)


def _compact_code(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).upper()


def _model_code_from_name(model_name: str) -> str:
    match = re.match(r"\s*([A-Z0-9/]+)\s*-", model_name or "", re.IGNORECASE)
    return _compact_code(match.group(1)) if match else ""


def _text_score(needle: str, haystack: str) -> float:
    needle_key = _match_key(needle)
    haystack_key = _match_key(haystack)
    if not needle_key or not haystack_key:
        return 0.0
    if needle_key == haystack_key:
        return 1.0
    if needle_key in haystack_key or haystack_key in needle_key:
        return 0.92
    needle_terms = set(needle_key.split())
    haystack_terms = set(haystack_key.split())
    overlap = len(needle_terms & haystack_terms) / max(len(needle_terms), 1)
    ratio = SequenceMatcher(None, needle_key, haystack_key).ratio()
    return max(ratio, overlap)


def _description_score(needle: str, haystack: str) -> Dict[str, Any]:
    scored = []
    for variant in _description_variants(needle):
        item = variant.copy()
        item["score"] = _text_score(variant["text"], haystack)
        if (
            item["score"] < MODULE_EXPANSION_MIN_SCORE
            and any(expansion.get("category") == "module" for expansion in item["expansions"])
        ):
            continue
        scored.append(item)

    return max(
        scored,
        key=lambda item: (
            item["score"],
            bool(item["expansions"]),
            -len(item["expansions"]),
        ),
        default={"text": _match_key(needle), "expansions": [], "score": 0.0},
    )


def _normalized_specification(value: Any) -> str:
    text = str(value or "").lower()
    text = text.replace("\u00b7", " ").replace("\u2022", " ")
    text = text.replace("\u2013", "-").replace("\u2014", "-")
    text = re.sub(r"(?<=\d),(?=\d)", ".", text)
    return re.sub(r"\s+", " ", text).strip()


def _has_explicit_torque_unit(text: str) -> bool:
    return any(
        re.search(unit_pattern, text or "", re.IGNORECASE)
        for _, unit_pattern in TORQUE_UNIT_PATTERNS
    )


def _measurement_values_match(
    target_values: List[float],
    found_values: List[float],
) -> bool:
    if len(target_values) != len(found_values):
        return False
    return all(
        abs(target - found) <= max(0.5, abs(target) * 0.03)
        for target, found in zip(target_values, found_values)
    )


def _is_alternate_unit_representation(
    previous: Dict[str, Any],
    current: Dict[str, Any],
    separator: str,
) -> bool:
    if previous["type"] != "torque" or current["type"] != "torque":
        return False
    if re.search(r"\+|\bthen\b|\bfollowed by\b", separator):
        return False
    if previous["unit"] == current["unit"]:
        return False
    return _measurement_values_match(previous["values_nm"], current["values_nm"])


def _parse_torque_specification(value: Any) -> Dict[str, Any]:
    text = _normalized_specification(value)
    angle_pattern = rf"(?P<value>{NUMBER_PATTERN})\s*(?:\u00b0|degrees?|deg)(?![a-z])"
    stages = []
    occupied_spans = []
    for unit, unit_pattern in TORQUE_UNIT_PATTERNS:
        pattern = re.compile(
            (
                rf"(?P<value>{NUMBER_PATTERN})"
                rf"(?:\s*(?:-|to)\s*(?P<range_end>{NUMBER_PATTERN}))?"
                rf"\s*(?P<unit>{unit_pattern})"
            ),
            re.IGNORECASE,
        )
        for match in pattern.finditer(text):
            raw_values = [float(match.group("value"))]
            if match.group("range_end") is not None:
                raw_values.append(float(match.group("range_end")))
            stages.append(
                {
                    "type": "torque",
                    "values": raw_values,
                    "values_nm": [
                        raw_value * TORQUE_CONVERSIONS_TO_NM[unit]
                        for raw_value in raw_values
                    ],
                    "unit": unit,
                    "span": match.span(),
                }
            )
            occupied_spans.append(match.span())

    for match in re.finditer(angle_pattern, text, re.IGNORECASE):
        stages.append(
            {
                "type": "angle",
                "value": float(match.group("value")),
                "span": match.span(),
            }
        )
        occupied_spans.append(match.span())

    if not any(stage["type"] == "torque" for stage in stages):
        shorthand_pattern = re.compile(
            (
                rf"(?P<value>{NUMBER_PATTERN})"
                rf"(?P<separator>\s*(?:\+|plus|then|followed\s+by)\s*)"
                rf"(?P<angle>{NUMBER_PATTERN})\s*(?:\u00b0|degrees?|deg)(?![a-z])"
            ),
            re.IGNORECASE,
        )
        for match in shorthand_pattern.finditer(text):
            stages.append(
                {
                    "type": "torque",
                    "values": [float(match.group("value"))],
                    "values_nm": [float(match.group("value"))],
                    "unit": "nm",
                    "span": match.span("value"),
                }
            )
            occupied_spans.append(match.span("value"))

    stages.sort(key=lambda stage: stage["span"][0])
    normalized_stages = []
    for stage in stages:
        if normalized_stages:
            previous = normalized_stages[-1]
            separator = text[previous["span"][1]:stage["span"][0]]
            if _is_alternate_unit_representation(previous, stage, separator):
                continue
        normalized_stages.append(stage)

    bare_numbers = []
    for match in re.finditer(NUMBER_PATTERN, text):
        if any(start <= match.start() and match.end() <= end for start, end in occupied_spans):
            continue
        bare_numbers.append(float(match.group()))

    bare_numeric_only = bool(re.fullmatch(NUMBER_PATTERN, text))
    has_unparsed_numeric_content = bool(bare_numbers and not bare_numeric_only)
    return {
        "text": text,
        "stages": normalized_stages,
        "bare_numbers": bare_numbers,
        "bare_numeric_only": bare_numeric_only,
        "has_unparsed_numeric_content": has_unparsed_numeric_content,
    }


def _torque_stages_match(target: Dict[str, Any], found: Dict[str, Any]) -> bool:
    if target["type"] != found["type"]:
        return False
    if target["type"] == "angle":
        return abs(target["value"] - found["value"]) < 0.01
    return _measurement_values_match(target["values_nm"], found["values_nm"])


def _torque_match(target: str, found: str) -> bool:
    target_spec = _parse_torque_specification(target)
    found_spec = _parse_torque_specification(found)

    if target_spec["has_unparsed_numeric_content"] or found_spec["has_unparsed_numeric_content"]:
        return False

    if target_spec["stages"]:
        if len(target_spec["stages"]) != len(found_spec["stages"]):
            return False
        return all(
            _torque_stages_match(target_stage, found_stage)
            for target_stage, found_stage in zip(
                target_spec["stages"],
                found_spec["stages"],
            )
        )

    if target_spec["bare_numeric_only"]:
        target_number = target_spec["bare_numbers"][0]
        if found_spec["bare_numeric_only"]:
            return abs(target_number - found_spec["bare_numbers"][0]) < 0.01
        if len(found_spec["stages"]) == 1 and found_spec["stages"][0]["type"] == "torque":
            found_values = found_spec["stages"][0]["values"]
            return len(found_values) == 1 and abs(target_number - found_values[0]) < 0.01
        return False

    target_key = _match_key(target)
    return bool(target_key and target_key in _match_key(found))


def _find_vehicle_versions(model_year: int, vehicle_family: str) -> List[Dict[str, Any]]:
    family_code = _compact_code(vehicle_family)
    matches = []
    priority_brand_codes = {
        "101": ["LANCIA", "RAM"],
        "103": ["RAM"],
        "105": ["LANCIA"],
        "106": ["LANCIA"],
        "107": ["LANCIA"],
        "109": ["LANCIA"],
        "110": ["FIAT"],
        "115": ["LANCIA"],
        "116": ["ALFA"],
        "118": ["FIAT"],
        "119": ["FIAT"],
        "121": ["FIAT"],
        "122": ["LANCIA"],
        "125": ["ALFA"],
        "130": ["ALFA"],
        "135": ["FIAT"],
        "136": ["ALFA"],
        "140": ["ALFA"],
        "141": ["FIAT"],
        "145": ["ALFA"],
        "146": ["FIAT"],
        "150": ["ABARTH", "FIAT"],
        "152": ["FIAT"],
        "157": ["LANCIA"],
        "158": ["FIAT"],
        "159": ["FIAT"],
        "160": ["FIAT"],
        "163": ["ALFA"],
        "164": ["ALFA"],
        "167": ["ALFA"],
        "169": ["FIAT"],
        "170": ["FIAT", "FIAT_PROFESSIONAL"],
        "171": ["FIAT"],
        "173": ["FIAT"],
        "174": ["ALFA"],
        "175": ["FIAT"],
        "176": ["FIAT"],
        "177": ["ALFA"],
        "178": ["FIAT"],
        "179": ["FIAT"],
        "180": ["LANCIA"],
        "181": ["CHRYSLER", "LANCIA"],
        "182": ["FIAT"],
        "183": ["FIAT"],
        "184": ["LANCIA"],
        "185": ["FIAT"],
        "186": ["FIAT"],
        "187": ["FIAT"],
        "188": ["FIAT"],
        "189": ["FIAT"],
        "190": ["ALFA"],
        "191": ["ALFA"],
        "192": ["FIAT"],
        "194": ["FIAT"],
        "195": ["FIAT"],
        "196": ["FIAT"],
        "197": ["FIAT"],
        "198": ["FIAT"],
        "199": ["ABARTH", "FIAT"],
        "222": ["FIAT", "FIAT_PROFESSIONAL"],
        "223": ["FIAT", "FIAT_PROFESSIONAL"],
        "225": ["FIAT_PROFESSIONAL"],
        "226": ["FIAT", "RAM"],
        "230": ["FIAT", "FIAT_PROFESSIONAL"],
        "235": ["FIAT_PROFESSIONAL"],
        "239": ["FIAT_PROFESSIONAL"],
        "241": ["FIAT_PROFESSIONAL"],
        "244": ["FIAT", "FIAT_PROFESSIONAL"],
        "245": ["FIAT", "FIAT_PROFESSIONAL"],
        "250": ["FIAT_PROFESSIONAL"],
        "255": ["FIAT", "FIAT_PROFESSIONAL"],
        "258": ["FIAT"],
        "263": ["FIAT_PROFESSIONAL"],
        "265": ["FIAT_PROFESSIONAL", "RAM"],
        "267": ["FIAT"],
        "271": ["FIAT_PROFESSIONAL"],
        "272": ["FIAT_PROFESSIONAL"],
        "276": ["FIAT_PROFESSIONAL"],
        "278": ["FIAT", "FIAT_PROFESSIONAL"],
        "281": ["FIAT", "FIAT_PROFESSIONAL", "RAM"],
        "285": ["FIAT_PROFESSIONAL"],
        "286": ["FIAT_PROFESSIONAL"],
        "287": ["FIAT_PROFESSIONAL"],
        "288": ["FIAT_PROFESSIONAL"],
        "289": ["FIAT_PROFESSIONAL"],
        "292": ["FIAT_PROFESSIONAL"],
        "294": ["FIAT_PROFESSIONAL"],
        "296": ["FIAT_PROFESSIONAL"],
        "298": ["FIAT_PROFESSIONAL"],
        "299": ["FIAT_PROFESSIONAL"],
        "300": ["FIAT"],
        "301": ["ABARTH", "FIAT"],
        "302": ["FIAT"],
        "305": ["FIAT"],
        "330": ["FIAT"],
        "332": ["ABARTH", "FIAT"],
        "334": ["FIAT"],
        "341": ["FIAT"],
        "343": ["FIAT"],
        "345": ["FIAT"],
        "348": ["ABARTH", "FIAT"],
        "354": ["FIAT"],
        "356": ["FIAT"],
        "357": ["FIAT"],
        "358": ["FIAT"],
        "359": ["FIAT"],
        "363": ["FIAT"],
        "365": ["ABARTH"],
        "372": ["FIAT"],
        "373": ["FIAT"],
        "374": ["FIAT"],
        "376": ["FIAT"],
        "402": ["CHRYSLER", "LANCIA"],
        "404": ["LANCIA"],
        "405": ["LANCIA"],
        "406": ["LANCIA"],
        "41": ["RAM"],
        "47": ["RAM"],
        "48": ["RAM"],
        "4C": ["ALFA"],
        "505": ["FIAT_PROFESSIONAL"],
        "508": ["FIAT_PROFESSIONAL"],
        "519": ["FIAT_PROFESSIONAL"],
        "530": ["FIAT_PROFESSIONAL"],
        "539": ["RAM"],
        "557": ["FIAT_PROFESSIONAL"],
        "57": ["FIAT", "RAM"],
        "578": ["FIAT"],
        "579": ["FIAT"],
        "580": ["FIAT_PROFESSIONAL"],
        "620": ["ALFA"],
        "630": ["ALFA"],
        "643": ["ALFA"],
        "67": ["RAM"],
        "A1": ["CHRYSLER", "DODGE"],
        "A3": ["CHRYSLER"],
        "A4": ["CHRYSLER"],
        "AA": ["CHRYSLER", "DODGE", "PLYMOUTH"],
        "AB": ["DODGE", "RAM"],
        "AC": ["CHRYSLER", "DODGE"],
        "AD": ["DODGE"],
        "AG": ["CHRYSLER", "DODGE", "PLYMOUTH"],
        "AJ": ["CHRYSLER"],
        "AN": ["DODGE"],
        "AP": ["DODGE", "PLYMOUTH"],
        "AS": ["CHRYSLER", "DODGE", "PLYMOUTH"],
        "AY": ["CHRYSLER"],
        "B1": ["JEEP"],
        "B2": ["DODGE", "EAGLE", "PLYMOUTH"],
        "B7": ["DODGE"],
        "B8": ["DODGE", "EAGLE", "PLYMOUTH"],
        "B9": ["DODGE", "EAGLE"],
        "BA": ["FIAT"],
        "BB": ["DODGE", "EAGLE"],
        "BD": ["EAGLE", "PLYMOUTH"],
        "BE": ["DODGE", "RAM"],
        "BF": ["FIAT"],
        "BG": ["FIAT"],
        "BK": ["DODGE", "FIAT"],
        "BQ": ["JEEP"],
        "BR": ["DODGE", "RAM"],
        "BT": ["DODGE"],
        "BU": ["JEEP"],
        "BV": ["JEEP"],
        "BW": ["DODGE"],
        "BX": ["DODGE"],
        "C": ["CITROEN"],
        "CS": ["CHRYSLER"],
        "D1": ["DODGE", "RAM"],
        "D2": ["RAM"],
        "DC": ["DODGE", "RAM"],
        "DD": ["RAM"],
        "DF": ["RAM"],
        "DH": ["DODGE", "RAM"],
        "DJ": ["RAM"],
        "DM": ["DODGE", "RAM"],
        "DN": ["DODGE"],
        "DP": ["RAM"],
        "DR": ["DODGE", "RAM"],
        "DS": ["RAM"],
        "DT": ["RAM"],
        "DV": ["DODGE"],
        "DX": ["DODGE", "RAM"],
        "EJ": ["JEEP"],
        "FB": ["FIAT"],
        "FD": ["FIAT"],
        "FF": ["FIAT"],
        "FG": ["FIAT"],
        "FJ": ["CHRYSLER", "DODGE", "EAGLE"],
        "GA": ["ALFA"],
        "GG": ["DODGE"],
        "GS": ["CHRYSLER"],
        "GU": ["ALFA"],
        "H1": ["JEEP"],
        "H6": ["JEEP"],
        "HB": ["DODGE"],
        "HG": ["CHRYSLER"],
        "J1": ["CHRYSLER"],
        "J2": ["CHRYSLER", "DODGE"],
        "J3": ["JEEP"],
        "J5": ["DODGE", "PEUGEOT"],
        "J6": ["JEEP"],
        "J8": ["JEEP"],
        "JA": ["CHRYSLER", "DODGE", "PLYMOUTH"],
        "JC": ["DODGE"],
        "JF": ["FIAT"],
        "JJ": ["JEEP"],
        "JK": ["JEEP"],
        "JL": ["JEEP"],
        "JR": ["CHRYSLER", "DODGE"],
        "JS": ["CHRYSLER", "DODGE", "LANCIA"],
        "JT": ["JEEP"],
        "JX": ["CHRYSLER"],
        "K1": ["JEEP"],
        "K4": ["JEEP"],
        "K8": ["JEEP"],
        "KA": ["DODGE"],
        "KB": ["RAM"],
        "KJ": ["JEEP"],
        "KK": ["JEEP"],
        "KL": ["JEEP"],
        "KM": ["JEEP"],
        "L2": ["CHRYSLER"],
        "LA": ["DODGE"],
        "LB": ["DODGE"],
        "LC": ["DODGE"],
        "LD": ["DODGE"],
        "LE": ["CHRYSLER"],
        "LH": ["CHRYSLER", "DODGE", "EAGLE"],
        "LX": ["CHRYSLER", "DODGE", "LANCIA"],
        "M1": ["JEEP"],
        "M4": ["JEEP"],
        "M6": ["JEEP"],
        "M7": ["JEEP"],
        "MJ": ["JEEP"],
        "MK": ["JEEP"],
        "MP": ["JEEP"],
        "MV": ["JEEP"],
        "MW": ["JEEP"],
        "N4": ["RAM"],
        "ND": ["DODGE"],
        "NS": ["CHRYSLER", "DODGE", "PLYMOUTH"],
        "NY": ["LANCIA"],
        "P1": ["DODGE"],
        "P2": ["CHRYSLER"],
        "P3": ["DODGE"],
        "P5": ["CHRYSLER"],
        "PD": ["DODGE"],
        "PF": ["DODGE"],
        "PG": ["CHRYSLER"],
        "PL": ["CHRYSLER", "DODGE", "PLYMOUTH"],
        "PM": ["DODGE"],
        "PR": ["CHRYSLER", "PLYMOUTH"],
        "PT": ["CHRYSLER"],
        "R2": ["CHRYSLER", "DODGE"],
        "RG": ["CHRYSLER"],
        "RM": ["VW"],
        "RS": ["CHRYSLER", "DODGE"],
        "RT": ["CHRYSLER", "DODGE", "LANCIA", "RAM"],
        "RU": ["CHRYSLER"],
        "SR": ["DODGE"],
        "ST": ["CHRYSLER", "DODGE"],
        "T1": ["JEEP"],
        "TJ": ["JEEP"],
        "UF": ["CHRYSLER"],
        "VA": ["DODGE"],
        "VB": ["DODGE"],
        "VF": ["RAM"],
        "VM": ["RAM"],
        "W1": ["JEEP"],
        "W2": ["JEEP"],
        "W3": ["JEEP"],
        "W4": ["JEEP"],
        "W7": ["JEEP"],
        "WD": ["DODGE"],
        "WG": ["JEEP"],
        "WH": ["JEEP"],
        "WJ": ["JEEP"],
        "WK": ["JEEP"],
        "WL": ["JEEP"],
        "WS": ["JEEP"],
        "X1": ["JEEP"],
        "X2": ["JEEP"],
        "XH": ["JEEP"],
        "XJ": ["JEEP"],
        "XK": ["JEEP"],
        "YJ": ["JEEP"],
        "ZB": ["DODGE"],
        "ZD": ["DODGE"],
        "ZG": ["JEEP"],
        "ZH": ["CHRYSLER"],
        "ZJ": ["JEEP"],
    }
    preferred_codes = priority_brand_codes.get(family_code, [])
    brand_items = list(config.BRAND_CODES.items())
    brand_items.sort(key=lambda item: (item[1] not in preferred_codes, item[0]))

    for brand_name, brand_code in brand_items:
        response = _get_json(
            "/connect/api/vehicle/models/categorized",
            params=config.get_model_request_params(brand_code),
        )
        for model in parser.extract_models(response):
            model_name = parser.get_model_name(model)
            model_code = _model_code_from_name(model_name)
            for version in parser.extract_versions(model):
                version_name = parser.get_version_name(version)
                version_model_code = _compact_code(version.get("modelCode") or model_code)
                if str(version_name).strip() != str(model_year):
                    continue
                if version_model_code != family_code and model_code != family_code:
                    continue
                matches.append(
                    {
                        "brand": brand_name,
                        "brand_code": brand_code,
                        "model": config.get_model_display_name(brand_code, model_name),
                        "model_code": version_model_code or model_code,
                        "version": version_name,
                        "model_version_id": parser.get_version_id(version),
                    }
                )
        if matches and brand_code in preferred_codes:
            break

    return matches


def _extract_engine_options(response: Dict[str, Any]) -> List[Dict[str, str]]:
    engines = []
    seen = set()
    for item in _walk(response):
        if not isinstance(item, dict):
            continue
        engine_name = _clean_text(
            item.get("engine")
            or item.get("engineName")
            or item.get("engineDescription")
            or item.get("motor")
            or item.get("powertrain")
        )
        engine_code = _clean_text(item.get("salesCode") or item.get("engineCode") or item.get("code"))
        engine_id = _clean_text(item.get("modelVersionEngineId"))
        if not engine_name:
            continue
        marker = (engine_name, engine_code, engine_id)
        if marker in seen:
            continue
        seen.add(marker)
        engines.append({"engine": engine_name, "engine_code": engine_code, "model_version_engine_id": engine_id})
    return engines


def _get_engine_options(vehicle: Dict[str, Any]) -> List[Dict[str, str]]:
    response = _get_json("/connect/api/vehicle/engines", params={"modelVersionId": vehicle["model_version_id"]})
    return _extract_engine_options(response)


def _find_engine(vehicle: Dict[str, Any], engine_code: str) -> Optional[Dict[str, str]]:
    wanted_code = _compact_code(engine_code)
    for engine in _get_engine_options(vehicle):
        if _compact_code(engine["engine_code"]) == wanted_code:
            return engine
    return None


def _get_service_book(model_version_engine_id: str) -> Optional[Dict[str, Any]]:
    response = _get_json(
        f"/connect/api/vehicle-tools/{CONFIG_LEVEL}/{model_version_engine_id}",
        params={"locale": config.MODEL_LOCALE, "nocache": str(int(time.time() * 1000))},
    )
    for book in response.get("books", []):
        if book.get("disciplineCode") == "service-info":
            return book
    return None


def _collect_torque_leaves(toc_response: Dict[str, Any]) -> List[Dict[str, str]]:
    leaves = []

    def visit(node: Dict[str, Any], path: List[str]) -> None:
        name = _clean_text(node.get("name"))
        current_path = path + ([name] if name else [])
        children = node.get("nodes") or []
        if children:
            for child in children:
                if isinstance(child, dict):
                    visit(child, current_path)
            return

        info_text = f"{name} {node.get('infoType', '')}"
        if "torque" not in info_text.lower():
            return
        content_link_id = node.get("contentLinkId")
        info_code = node.get("infoCode")
        if content_link_id and info_code:
            leaves.append(
                {
                    "path": " / ".join(current_path),
                    "name": name,
                    "info_type": _clean_text(node.get("infoType")),
                    "content_link_id": content_link_id,
                    "info_code": info_code,
                }
            )

    for root in toc_response.get("nodes", []):
        if isinstance(root, dict):
            visit(root, [])
    return leaves


def _rank_leaves(leaves: List[Dict[str, str]], vsc_name: str) -> List[Dict[str, str]]:
    ranked = []
    for leaf in leaves:
        score = _description_score(vsc_name, leaf["path"])["score"]
        item = leaf.copy()
        item["vsc_score"] = score
        ranked.append(item)
    return sorted(ranked, key=lambda item: item["vsc_score"], reverse=True)


def _column_key(column: Any) -> str:
    if isinstance(column, tuple):
        parts = []
        for part in column:
            text = _clean_text(part)
            if not text or text.lower() == "nan" or text.lower().startswith("unnamed:"):
                continue
            parts.append(text)
        return _match_key(" ".join(parts))
    return _match_key(column)


def _is_empty_spec_value(value: Any) -> bool:
    text = _clean_text(value)
    key = _match_key(text)
    return not text or key in {"nan", "none"} or bool(re.fullmatch(r"[-\u2013\u2014]+", text))


def _format_unit_specification(value: Any, unit: str) -> str:
    text = _clean_text(value)
    if re.fullmatch(r"[-+]?\d+\.0", text):
        text = text[:-2]
    if not text or _has_explicit_torque_unit(text):
        return text

    unit_label = TORQUE_UNIT_DISPLAY[unit]
    staged = re.match(
        rf"^(?P<value>{NUMBER_PATTERN})(?P<tail>\s*(?:\+|plus|then|followed\s+by)\s*.*)$",
        text,
        re.IGNORECASE,
    )
    if staged:
        return f"{staged.group('value')} {unit_label}{staged.group('tail')}"
    return f"{text} {unit_label}"


def _unit_column_for_key(column_key: str) -> Optional[str]:
    return TORQUE_TABLE_UNIT_COLUMNS.get(column_key)


def _unit_columns_from_table(table: pd.DataFrame) -> List[Tuple[Any, str]]:
    unit_columns = []
    for column in table.columns:
        unit = _unit_column_for_key(_column_key(column))
        if unit:
            unit_columns.append((column, unit))
    return unit_columns


def _positional_unit_columns(table: pd.DataFrame, start_index: int = 1) -> List[Tuple[Any, str]]:
    columns = list(table.columns)
    positional_units = ["nm", "ft_lb", "in_lb"]
    return [
        (columns[index], unit)
        for index, unit in enumerate(positional_units, start=start_index)
        if index < len(columns)
    ]


def _row_unit_specification(row: pd.Series, unit_columns: List[Tuple[Any, str]]) -> str:
    parts = []
    for column, unit in unit_columns:
        value = row.get(column)
        if _is_empty_spec_value(value):
            continue
        parts.append(_format_unit_specification(value, unit))
    return " / ".join(parts)


def _extract_torque_rows(content_html: str, leaf: Dict[str, str]) -> List[Dict[str, str]]:
    rows = []
    try:
        tables = pd.read_html(StringIO(content_html))
    except ValueError:
        tables = []
    # If pandas found no tables in the provided HTML, try fallback fetches
    if not tables:
        # Look for direct raw content links in the shell HTML and try to fetch them
        hrefs = re.findall(r'href=["\']([^"\']*connect/api/content/raw/[^"\']*)', content_html, flags=re.I)
        for href in hrefs:
            href = href.replace('&amp;', '&')
            parsed = urlparse(href)
            if parsed.path:
                path = parsed.path
            else:
                path = href
            params = {k: v[0] for k, v in parse_qs(parsed.query).items()} if parsed.query else {}
            # Ensure auth token present in params if available
            auth_token = config.get_auth_token()
            if auth_token and 'X-Auth-Token' not in params:
                params['X-Auth-Token'] = auth_token
            try:
                raw_html = _get_text(path, params=params)
            except requests.RequestException:
                raw_html = None
            if raw_html:
                try:
                    tables = pd.read_html(StringIO(raw_html))
                except ValueError:
                    tables = []
                if tables:
                    content_html = raw_html
                    break

    # If still no tables, try plugin iframe endpoint as a last resort
    if not tables:
        try:
            metadata = _get_content_metadata(leaf, '')
        except Exception:
            metadata = {}
        try:
            iframe_params = _plugin_iframe_params(leaf, '', metadata)
            iframe_html = _get_text('/web/secure/api/plugin/iframe', params=iframe_params)
            try:
                tables = pd.read_html(StringIO(iframe_html))
                content_html = iframe_html
            except ValueError:
                tables = []
        except Exception:
            tables = []

    for table in tables:
        normalized_columns = [_column_key(column).upper() for column in table.columns]
        description_column = None
        specification_column = None
        comment_column = None

        if "DESCRIPTION" in normalized_columns:
            description_column = table.columns[normalized_columns.index("DESCRIPTION")]
        if "SPECIFICATION" in normalized_columns:
            specification_column = table.columns[normalized_columns.index("SPECIFICATION")]
        if "COMMENT" in normalized_columns:
            comment_column = table.columns[normalized_columns.index("COMMENT")]

        unit_columns = _unit_columns_from_table(table)
        if description_column is None and specification_column is None and not unit_columns and len(table.columns) >= 2:
            description_column = table.columns[0]
            unit_columns = _positional_unit_columns(table)
            if len(table.columns) > 4:
                comment_column = table.columns[4]

        if description_column is not None and specification_column is None and not unit_columns:
            description_index = list(table.columns).index(description_column)
            unit_columns = _positional_unit_columns(table, description_index + 1)
            if comment_column in {column for column, _ in unit_columns}:
                unit_columns = [item for item in unit_columns if item[0] != comment_column]

        if description_column is None or (specification_column is None and not unit_columns):
            continue

        for _, row in table.iterrows():
            description = _clean_text(row.get(description_column))
            if not description or description.lower() == "nan" or _column_key(description) == "description":
                continue

            if specification_column is not None:
                specification = _clean_text(row.get(specification_column))
            else:
                specification = _row_unit_specification(row, unit_columns)
            if _is_empty_spec_value(specification):
                continue

            rows.append(
                {
                    "page": leaf["path"],
                    "description": description,
                    "specification": specification,
                    "comment": (
                        ""
                        if comment_column is None or _is_empty_spec_value(row.get(comment_column))
                        else _clean_text(row.get(comment_column))
                    ),
                }
            )
    return rows


def _score_torque_row(
    row: Dict[str, Any],
    leaf: Dict[str, Any],
    description: str,
    target_torque: str,
) -> Dict[str, Any]:
    description_result = (
        _description_score(description, row["description"])
        if description.strip()
        else {"text": "", "expansions": [], "score": 0.0}
    )
    description_score = description_result["score"]
    torque_matches = _torque_match(target_torque, row["specification"])
    torque_score = 1.0 if torque_matches else _text_score(target_torque, row["specification"])
    vsc_score = float(leaf.get("vsc_score", 0.0))

    if description.strip():
        score = (0.58 * description_score) + (0.30 * torque_score) + (0.12 * vsc_score)
    else:
        score = (0.78 * torque_score) + (0.22 * vsc_score)

    enriched = row.copy()
    enriched["vsc_score"] = vsc_score
    enriched["description_score"] = description_score
    enriched["matched_description_input"] = description_result["text"]
    enriched["shortcut_expansions"] = description_result["expansions"]
    enriched["torque_score"] = torque_score
    enriched["torque_match"] = torque_matches
    enriched["score"] = score
    enriched["confidence"] = round(score * 100, 1)
    return enriched


def _ambiguous_expansion_signature(candidate: Dict[str, Any]) -> Tuple[Tuple[str, str], ...]:
    return tuple(
        sorted(
            (item["shortcut"], item["meaning"])
            for item in candidate.get("shortcut_expansions", [])
            if item.get("ambiguous")
        )
    )


def _find_ambiguous_competitor(
    best: Dict[str, Any],
    ranked_candidates: List[Dict[str, Any]],
) -> Optional[Dict[str, Any]]:
    best_signature = _ambiguous_expansion_signature(best)
    if not best_signature:
        return None

    for candidate in ranked_candidates[1:]:
        candidate_signature = _ambiguous_expansion_signature(candidate)
        if not candidate_signature or candidate_signature == best_signature:
            continue
        if not candidate.get("torque_match") or candidate.get("description_score", 0.0) < 0.65:
            continue
        if best["score"] - candidate["score"] <= 0.03:
            return candidate
    return None


def _description_has_ambiguous_shortcuts(description: str) -> bool:
    return any(
        expansion.get("ambiguous")
        for variant in _description_variants(description)
        for expansion in variant.get("expansions", [])
    )


def _initial_torque_leaves(
    leaves: List[Dict[str, Any]],
    vsc_name: str,
) -> List[Dict[str, Any]]:
    strong_vsc_leaves = [leaf for leaf in leaves if leaf["vsc_score"] >= 0.35]
    if vsc_name.strip():
        return strong_vsc_leaves[:12] if strong_vsc_leaves else leaves[:25]
    return leaves[:35]


def _has_decisive_candidate(
    candidates: List[Dict[str, Any]],
    description: str,
) -> bool:
    if not description.strip() or _description_has_ambiguous_shortcuts(description):
        return False
    return any(
        candidate["description_score"] >= 0.95
        and candidate["torque_match"]
        and not _ambiguous_expansion_signature(candidate)
        for candidate in candidates
    )


def _build_engine_targets(
    vehicles: List[Dict[str, Any]],
    engine_code: str,
) -> tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    targets = []
    checked_engines = []
    wanted_code = _compact_code(engine_code)
    seen = set()

    for vehicle in vehicles:
        engines = _get_engine_options(vehicle)
        matched_engines = []
        for engine in engines:
            if wanted_code and _compact_code(engine["engine_code"]) != wanted_code:
                continue
            marker = (vehicle["model_version_id"], engine.get("model_version_engine_id"))
            if marker in seen:
                continue
            seen.add(marker)
            matched_engines.append(engine)
            targets.append({"vehicle": vehicle, "engine": engine})

        checked_engines.append(
            {
                "vehicle": vehicle,
                "engine_found": bool(matched_engines),
                "engine_count": len(engines),
            }
        )

    return targets, checked_engines


def verify_torque(
    model_year: int,
    vehicle_family: str,
    engine_code: str,
    vsc_name: str,
    description: str,
    target_torque: str,
) -> Dict[str, Any]:
    """Verify a torque row against Service Library for one vehicle/engine."""
    vehicles = _find_vehicle_versions(model_year, vehicle_family)
    if not vehicles:
        return {
            "vehicle_match": False,
            "engine_match": False,
            "vsc_match": False,
            "description_match": False,
            "torque_match": False,
            "message": "No vehicle found for this model year and VEH FAM.",
            "candidates": [],
        }

    engine_code_was_provided = bool(engine_code.strip())
    engine_targets, checked_engines = _build_engine_targets(vehicles, engine_code)

    if not engine_targets:
        return {
            "vehicle_match": True,
            "engine_match": False,
            "vsc_match": False,
            "description_match": False,
            "torque_match": False,
            "status": "Needs review",
            "confidence": 0,
            "message": "Vehicle found, but the engine code was not found for that vehicle.",
            "vehicles": vehicles,
            "checked_engines": checked_engines,
            "candidates": [],
        }

    all_candidates = []
    selected_vehicle = engine_targets[0]["vehicle"]
    selected_engine = engine_targets[0]["engine"]
    pages_found = 0
    pages_checked = 0
    readable_pages = 0
    unreadable_pages = 0
    content_errors = []
    engine_books_checked = 0
    missing_books = 0

    for target in engine_targets:
        vehicle = target["vehicle"]
        engine = target["engine"]
        service_book = _get_service_book(engine["model_version_engine_id"])
        if not service_book:
            missing_books += 1
            continue

        engine_books_checked += 1
        toc = _get_json(
            f"/connect/api/toc/{service_book['modelVersionBookId']}/{CONFIG_LEVEL}/{engine['model_version_engine_id']}",
            params={"locale": config.MODEL_LOCALE, "nocache": str(int(time.time() * 1000))},
        )
        leaves = _rank_leaves(_collect_torque_leaves(toc), vsc_name)
        pages_found = max(pages_found, len(leaves))
        if not leaves:
            continue

        initial_leaves = _initial_torque_leaves(leaves, vsc_name)
        initial_ids = {leaf["content_link_id"] for leaf in initial_leaves}
        remaining_leaves = [
            leaf for leaf in leaves if leaf["content_link_id"] not in initial_ids
        ]
        target_candidates = []

        # Per-target debug info to help diagnose skipped pages in production
        if "search_debug" not in locals():
            search_debug = []
        target_debug = {
            "vehicle": vehicle,
            "engine": engine,
            "pages_found": len(leaves),
            "initial_leaves_count": len(initial_leaves),
            "remaining_leaves_count": len(remaining_leaves),
            "initial_checked": 0,
            "remaining_checked": 0,
            "decisive_after_initial": False,
        }

        def search_pages(pages: List[Dict[str, Any]]) -> None:
            nonlocal pages_checked, readable_pages, unreadable_pages
            for leaf in pages:
                pages_checked += 1
                try:
                    html = _get_torque_content(leaf, engine["model_version_engine_id"])
                except requests.RequestException as exc:
                    unreadable_pages += 1
                    if len(content_errors) < 10:
                        response = exc.response
                        content_errors.append(
                            {
                                "page": leaf["path"],
                                "content_link_id": leaf["content_link_id"],
                                "status": response.status_code if response is not None else None,
                                "error": exc.__class__.__name__,
                            }
                        )
                    continue
                readable_pages += 1
                for row in _extract_torque_rows(html, leaf):
                    candidate = _score_torque_row(row, leaf, description, target_torque)
                    candidate["vehicle"] = vehicle
                    candidate["engine"] = engine
                    candidate["engine_code_provided"] = engine_code_was_provided
                    target_candidates.append(candidate)
                    all_candidates.append(candidate)

        # Search initial leaves first and record how many pages were checked
        before = pages_checked
        search_pages(initial_leaves)
        target_debug["initial_checked"] = pages_checked - before

        # Decide whether to search remaining leaves
        decisive = _has_decisive_candidate(target_candidates, description)
        target_debug["decisive_after_initial"] = bool(decisive)
        if remaining_leaves and not decisive:
            before_rem = pages_checked
            search_pages(remaining_leaves)
            target_debug["remaining_checked"] = pages_checked - before_rem

        search_debug.append(target_debug)

    if not engine_books_checked:
        return {
            "vehicle_match": True,
            "engine_match": bool(engine_targets),
            "vsc_match": False,
            "description_match": False,
            "torque_match": False,
            "status": "Needs review",
            "confidence": 0,
            "message": "Vehicle and engine found, but Service Information book was not available.",
            "vehicle": selected_vehicle,
            "engine": selected_engine,
            "engines_checked": len(engine_targets),
            "missing_books": missing_books,
            "candidates": [],
        }

    ranked_candidates = sorted(all_candidates, key=lambda row: row["score"], reverse=True)
    candidates = ranked_candidates[:5]
    best = candidates[0] if candidates else None
    if not best:
        incomplete = unreadable_pages > 0
        return {
            "vehicle_match": True,
            "engine_match": engine_code_was_provided,
            "vsc_match": False,
            "description_match": False,
            "torque_match": False,
            "status": "Incomplete" if incomplete else "Not found",
            "confidence": 0,
            "message": (
                "No matching torque rows were found in readable pages, but some "
                "Service Library pages could not be checked."
                if incomplete
                else "No matching torque rows were found."
            ),
            "vehicle": selected_vehicle,
            "engine": selected_engine,
            "engines_checked": len(engine_targets),
            "candidates": [],
            "torque_pages_checked": pages_checked,
            "torque_pages_found": pages_found,
            "readable_torque_pages": readable_pages,
            "unreadable_torque_pages": unreadable_pages,
            "content_errors": content_errors,
            "search_debug": search_debug if 'search_debug' in locals() else [],
        }

    description_match = bool(description.strip() and best["description_score"] >= 0.65)
    torque_match = bool(best["torque_match"])
    ambiguous_competitor = _find_ambiguous_competitor(best, ranked_candidates)
    if torque_match and description_match and (engine_code_was_provided or len(engine_targets) == 1):
        status = "Verified"
    elif torque_match and (description_match or not description.strip()):
        status = "Probable match"
    elif best["confidence"] >= 55:
        status = "Needs review"
    else:
        status = "Not found"
    if ambiguous_competitor and status in {"Verified", "Probable match"}:
        status = "Needs review"
    if unreadable_pages and status in {"Verified", "Probable match"}:
        status = "Needs review"

    return {
        "vehicle_match": True,
        "engine_match": engine_code_was_provided and bool(engine_targets),
        "engine_code_provided": engine_code_was_provided,
        "vsc_match": bool(vsc_name.strip() and best["vsc_score"] >= 0.35),
        "description_match": description_match,
        "torque_match": torque_match,
        "shortcut_ambiguous": bool(ambiguous_competitor),
        "ambiguous_competitor": ambiguous_competitor,
        "status": status,
        "confidence": best["confidence"],
        "message": (
            "Verification completed with unreadable Service Library pages."
            if unreadable_pages
            else "Verification completed."
        ),
        "vehicle": best["vehicle"],
        "engine": best["engine"],
        "engines_checked": len(engine_targets),
        "missing_books": missing_books,
        "best": best,
        "candidates": candidates,
        "torque_pages_checked": pages_checked,
        "torque_pages_found": pages_found,
        "readable_torque_pages": readable_pages,
        "unreadable_torque_pages": unreadable_pages,
        "content_errors": content_errors,
        "search_debug": search_debug if 'search_debug' in locals() else [],
    }
